import openpyxl
from openpyxl.styles import Font, PatternFill

from openpyxl.comments import Comment

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B1 Match Data"

    # Define headers and their explanations
    headers_info = {
        "Unique ID": "Unique identifier for the pair (e.g. B1_001). used for tracking",
        "English Word": "The English translation. This will appear on the 'Text' card.",
        "French Word": "The French word. This text triggers the TTS (Text-to-Speech) for the 'Audio' card.",
        "Image URL": "OPTIONAL. URL to an image. If provided, can be used for Image-Text matching mode.",
        "Audio URL": "OPTIONAL. URL to a custom audio file. If empty, the app uses auto-generated TTS from 'French Word'.",
        "Level": "CEFR Level (e.g. B1). Used for filtering.",
        "Category": "Topic category (e.g. Animals, Work). Used for grouping."
    }

    headers = list(headers_info.keys())

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        
        # Add Comment
        explanation = headers_info[header]
        cell.comment = Comment(explanation, "System")
        
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Add some sample rows/instructions
    samples = [
        ["B1_001", "Dog", "Chien", "https://example.com/dog.jpg", "", "B1", "Animals"],
        ["B1_002", "Cat", "Chat", "https://example.com/cat.jpg", "", "B1", "Animals"],
    ]

    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    filename = "MatchPairsB1_Template.xlsx"
    wb.save(filename)
    print(f"Created {filename} successfully.")

if __name__ == "__main__":
    create_template()
